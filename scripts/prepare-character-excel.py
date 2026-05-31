import argparse
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKSPACE = ROOT.parent
SOURCE = WORKSPACE / "Relink" / "imports" / "maygi-damage-calculator.xlsx"
REPORT_DIR = ROOT / "reports"


def main():
    parser = argparse.ArgumentParser(description="Prepare a Google-Sheet-recalc input workbook for one character.")
    parser.add_argument("character", help="角色繁中名稱，例如：蘭斯洛特")
    parser.add_argument("--output", help="輸出 xlsx 路徑；預設在 reports 內")
    args = parser.parse_args()

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = args.character.replace("/", "-").replace("\\", "-").replace(" ", "")
    output = Path(args.output) if args.output else REPORT_DIR / f"{safe_name}-input.xlsx"

    wb = load_workbook(SOURCE)
    ws = wb["Calculator"]
    ws["B5"] = args.character
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(output)
    print(output)


if __name__ == "__main__":
    main()
