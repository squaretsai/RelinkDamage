import argparse
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKSPACE = ROOT.parent
SOURCE = WORKSPACE / "Relink" / "imports" / "maygi-damage-calculator.xlsx"
REPORT_DIR = ROOT / "reports"

BUILD = [
    ("傷害上限", 12, "攻擊力"),
    ("暴擊機率", 9, "暴擊傷害"),
    ("渾身", 14, "背水"),
    ("暴君", 10, "捨身"),
    ("技能傷害", 8, "連技加成"),
    ("連技終擊", 11, "蓄力攻擊"),
    ("集中砲火", 6, "投擲"),
    ("弱點攻擊", 13, "蓄力加速"),
    ("先制", 7, "逆襲閃避"),
    ("一線之隔", 5, "有利屬性轉換"),
    ("阿爾法", 16, "貝塔"),
    ("專屬", 15, "戰氣"),
]

WEAPON_TRAITS = [
    ("伽馬", 15),
    ("追擊", 7),
    ("狂戰士", 10),
]

LIMIT_BREAK = {
    "B16": 777,
    "B17": 0.18,
    "B18": 0.19,
    "B19": 0.07,
    "B20": 0.11,
    "B21": 0.13,
    "B22": 0.22,
}

OTHER_INPUTS = {
    "F26": 2,
    "F27": 0.12,
    "F28": 0.08,
    "F29": True,
    "F30": 0.37,
    "F31": True,
    "F32": True,
    "F33": True,
}


def safe_name(value: str) -> str:
    return value.replace("/", "-").replace("\\", "-").replace(" ", "")


def main():
    parser = argparse.ArgumentParser(description="Prepare one high-coverage stress workbook for a character.")
    parser.add_argument("character", help="角色繁中名稱，例如：蘭斯洛特")
    parser.add_argument("--output", help="輸出 xlsx 路徑；預設在 reports 內")
    args = parser.parse_args()

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    output = Path(args.output) if args.output else REPORT_DIR / f"{safe_name(args.character)}-stress-input.xlsx"

    wb = load_workbook(SOURCE)
    ws = wb["Calculator"]
    ws["B5"] = args.character

    for index, (main, level, sub) in enumerate(BUILD, start=7):
        ws[f"E{index}"] = main
        ws[f"F{index}"] = level
        ws[f"H{index}"] = sub

    for index, (trait, level) in enumerate(WEAPON_TRAITS, start=21):
        ws[f"E{index}"] = trait
        ws[f"F{index}"] = level

    ws["H22"] = True
    ws["H24"] = True

    for cell, value in LIMIT_BREAK.items():
        ws[cell] = value
    for cell, value in OTHER_INPUTS.items():
        ws[cell] = value

    # Keep generic character extras active when present. Some rows are formulas
    # that become N/A for unrelated characters after Google Sheet recalculates.
    for cell in ("B26", "B27", "B28"):
        ws[cell] = True

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(output)
    print(output)


if __name__ == "__main__":
    main()
