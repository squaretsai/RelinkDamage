from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKSPACE = ROOT.parent
SOURCE = WORKSPACE / "Relink" / "imports" / "maygi-damage-calculator.xlsx"
OUTPUT = ROOT / "reports" / "zeta-stress-input.xlsx"


BUILD = [
    ("傷害上限", 12, "攻擊力"),
    ("暴擊機率", 9, "暴擊傷害"),
    ("渾身", 14, "背水"),
    ("暴君", 10, "捨身"),
    ("技能傷害", 8, "連技加成"),
    ("連技終擊", 11, "蓄力攻擊"),
    ("集中砲火", 6, "投擲"),
    ("弱點攻擊", 13, "蓄力加速"),
    ("先制", 7, "逆襲閃避"),
    ("一線之隔", 5, "有利屬性轉換"),
    ("阿爾法", 16, "貝塔"),
    ("專屬", 15, "戰氣"),
]

WEAPON_TRAITS = [
    ("伽馬", 15),
    ("追擊", 7),
    ("狂戰士", 10),
]

LIMIT_BREAK = {
    "B16": 777,
    "B17": 0.18,
    "B18": 0.19,
    "B19": 0.07,
    "B20": 0.11,
    "B21": 0.13,
    "B22": 0.22,
}

OTHER_INPUTS = {
    "F26": 2,
    "F27": 0.12,
    "F28": 0.08,
    "F29": True,
    "F30": 0.37,
    "F31": True,
    "F32": True,
    "F33": True,
}

CHARACTER_EXTRAS = {
    "B26": True,
    "B28": True,
}


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(SOURCE)
    ws = wb["Calculator"]

    ws["B5"] = "瑟塔"

    for index, (main, level, sub) in enumerate(BUILD, start=7):
        ws[f"E{index}"] = main
        ws[f"F{index}"] = level
        ws[f"H{index}"] = sub

    for index, (trait, level) in enumerate(WEAPON_TRAITS, start=21):
        ws[f"E{index}"] = trait
        ws[f"F{index}"] = level

    ws["H22"] = True
    ws["H24"] = True

    for cell, value in LIMIT_BREAK.items():
        ws[cell] = value
    for cell, value in OTHER_INPUTS.items():
        ws[cell] = value
    for cell, value in CHARACTER_EXTRAS.items():
        ws[cell] = value

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
